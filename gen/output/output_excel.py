'''
This module exports the worload date to an excel file.
The template RCD_template.xslx is used as a base for export.
'''
from os import makedirs, path
from typing import List
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, PatternFill, Side

from _version import __version__
from model.config import Config
from model.resource_requirements import PVC, WL
from model.summary import Summary, ConfigMap, Secret


START_ROW = 14
CM_SECRET_START_ROW = 6
BLUE_FILL = PatternFill(fill_type='solid', start_color='dce6f2', end_color='dce6f2')
WHITE_FILL = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
RCD_EXCEL_TEMPLATE_PATH = './RCD_template.xlsx'
RCD_EIC_EXCEL_TEMPLATE_PATH = './RCD_EIC_template.xlsx'
EIC_VARIANT = 'eric-eic-integration-fixed-size-production-values'


class Storage:
    '''
    Storage class to export resource information to an Excel workbook.
    '''
    def __init__(self, config: Config):
        '''
        Initialise Storage
        '''
        self.config = config
        self.folder_path = config.output_folder
        makedirs(self.folder_path, exist_ok=True)

    def generate_workbook(self, summary: Summary, variant: str, version: str,
              eic_product = False, eic_app_name = None):
        '''
        Store all resource requirements in Excel workbook
        '''
        if eic_product:
            template_path = RCD_EIC_EXCEL_TEMPLATE_PATH
            page_heading = f"Deployment Type : {EIC_VARIANT}   Version: {version}"
        else:
            template_path = RCD_EXCEL_TEMPLATE_PATH
            page_heading = \
                f"Deployment Type : {self.config.variants[variant][0]}   Version: {version}"

        template = path.join(path.dirname(
            path.abspath(__file__)), template_path)
        workbook = load_workbook(filename=template)
        sgr_sheet = workbook['Workloads']
        sgr_sheet.cell(column=1, row=1, value=page_heading)
        Storage.write_workloads(sgr_sheet, summary.workloads)

        sr_sheet = workbook['PVC Storage']
        sr_sheet.cell(column=1, row=1, value=page_heading)
        Storage.write_pvcs(sr_sheet, summary.pvcs)

        return workbook

    @staticmethod
    def apply_style(cell: Cell, fill: PatternFill):
        '''
        Apply scecified Fill style to a Cell
        '''
        cell.fill = fill
        side = Side(style='thin')
        cell.border = Border(top=side, left=side, right=side, bottom=side)
        if (isinstance(cell.value, str) and not cell.value.startswith('=')) \
            or isinstance(cell.value, bool):
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='right')

    @staticmethod
    def update_worksheet_data(worksheet, row, res_ws_data, fill: PatternFill = BLUE_FILL):
        '''
        Populate and apply style to specified row
        '''
        for data in res_ws_data:
            Storage.apply_style(worksheet.cell(column=data[0], row=row, value=data[1]), fill)

    @staticmethod
    def write_workloads(worksheet, workloads: List[WL]):
        '''
        Prepare data in Workloads worksheet
        '''
        row = START_ROW
        for workload in workloads:
            wl_data = [[1, workload['name']],
                       [2, ''],
                       [3, workload['app_enabled']],
                       [4, workload['kind']],
                       [5, f"=SUM(E{row+1}:E{row + len(workload['containers'])})"],
                       [6, f"=SUM(F{row+1}:F{row + len(workload['containers'])})"],
                       [7, f"=SUM(G{row+1}:G{row + len(workload['containers'])})"],
                       [8, f"=SUM(H{row+1}:H{row + len(workload['containers'])})"],
                       [9, f"=SUM(I{row+1}:I{row + len(workload['containers'])})"],
                       [10, f"=SUM(J{row+1}:J{row + len(workload['containers'])})"],
                       [11, workload['replicas'] if workload['replicas'] > 0 else ''],
                       [12, workload['chart']]
                       ]
            Storage.update_worksheet_data(worksheet, row, wl_data)
            row += 1
            for container in workload['containers']:
                c_data = [[1, ''],
                          [2, container['name']],
                          [3, ''],
                          [4, 'Container'],
                          [5, container['cpu_req']],
                          [6, container['cpu_lim']],
                          [7, container['mem_req']],
                          [8, container['mem_lim']],
                          [9, container['eps_req']*1024],
                          [10, container['eps_lim']*1024],
                          [11, ''],
                          [12, '']
                          ]
                Storage.update_worksheet_data(worksheet, row, c_data, WHITE_FILL)
                row += 1

    @staticmethod
    def write_pvcs(worksheet, pvcs: List[PVC]):
        '''
        Prepare data in Storage worksheet
        '''
        row = START_ROW
        for pvc in pvcs:
            pvc_data = [[1, pvc['name']],
                        [2, pvc['app_enabled']],
                        [3, pvc['instances'] if hasattr(pvc, 'instances') else 1],
                        [4, pvc['type']],
                        [5, pvc['size']],
                        [6, pvc['fullBackup']],
                        [7, pvc['rollback']],
                        [8, pvc['appName']]
                        ]
            Storage.update_worksheet_data(worksheet, row, pvc_data)
            row += 1

    @staticmethod
    def write_config_maps(worksheet, config_maps: List[ConfigMap]):
        '''
        Prepare data in Config Maps worksheet
        '''
        row = CM_SECRET_START_ROW
        for config_map in config_maps:
            config_map_data = [[1, config_map.name],
                        [2, config_map.app_enabled]
                        ]
            Storage.update_worksheet_data(worksheet, row, config_map_data)
            row += 1

    @staticmethod
    def write_secrets(worksheet, secrets: List[Secret]):
        '''
        Prepare data in Secrets worksheet
        '''
        row = CM_SECRET_START_ROW
        for secret in secrets:
            secret_data = [[1, secret.name],
                        [2, secret.app_enabled]
                        ]
            Storage.update_worksheet_data(worksheet, row, secret_data)
            row += 1
